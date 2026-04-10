"""
Parser del Excel de Quenti — versión Python "de producción".

Misma lógica que el archivo b370-manager/tests/test_parser_against_real_xlsx.py
pero en módulo importable, no en script de prueba.

Cobertura validada contra CUENTI INVENTARIO 6 ABRIL.xlsx: 1.077 / 1.090 = 98,8%.
Las 13 filas restantes son `BUSO DE ARQUERO/<6..18>` (uniformes infantiles que
Beto va a subir más adelante, no en este flujo).

Decisiones acordadas con Juanjo (no cambiar sin pedir):
- Tallas válidas: XS, S, M, L, XL, 2XL, 3XL, 4XL, 5XL, 6XL (XXL se normaliza a 2XL).
- Calidad: 'tipo_original' | 'version_fan' | '1.1' | None.
- Acabado: 'con_parches' solo si el token aparece explícito; ausencia → 'sin_parches'.
- RETRO queda en el nombre base, NO se extrae como atributo.
- Prefijos aceptados: CAMISETA, CAMISETA DE, CAMISETA DEL, BUSO, BUSO DE, BUSO DEL.
- 1,1 (coma colombiana) == 1.1.
"""

from __future__ import annotations

import re
from collections import defaultdict
from typing import Iterable, Iterator, Optional

from openpyxl import load_workbook

VALID_SIZES = ("XS", "S", "M", "L", "XL", "2XL", "3XL", "4XL", "5XL", "6XL")

CALIDAD_FAN      = "version_fan"
CALIDAD_ORIGINAL = "tipo_original"
CALIDAD_UNO_UNO  = "1.1"

ACABADO_CON = "con_parches"
ACABADO_SIN = "sin_parches"

TIPO_CAMISETA = "camiseta"
TIPO_BUSO     = "buso"

# Tallas que sabemos que NO existen aún en el atributo Tallas de WooCommerce.
# El CLI debe pedir confirmación a Juanjo antes de crearlas en la primera corrida.
SIZES_NOT_YET_IN_WC = ("XS", "3XL", "4XL", "5XL", "6XL")

_PREFIX_RE = re.compile(r"^(?P<head>CAMISETA|BUSO)(?:\s+DEL?)?\s+(?P<rest>.+)$")
_SIZE_TAIL_RE = re.compile(r"^(?P<body>.+?)\s*/\s*(?P<size>[A-Z0-9]+)\s*$", re.UNICODE)
_PARCHES_CON_RE = re.compile(r"\bCON\s+PARCHES\b")
_PARCHES_SIN_RE = re.compile(r"\bSIN\s+PARCHES\b")
_CALIDAD_PATTERNS = (
    (CALIDAD_ORIGINAL, re.compile(r"\bTIPO\s+ORIGINAL\b")),
    (CALIDAD_UNO_UNO,  re.compile(r"(?<![\d\.,])1[.,]1(?![\d])")),
    (CALIDAD_FAN,      re.compile(r"\bFAN\b")),
)

_TRANS = str.maketrans("ÁÉÍÓÚÜÑ", "AEIOUUN")


def normalize(s: object) -> str:
    if s is None:
        return ""
    s = str(s).lstrip("\ufeff").upper().translate(_TRANS)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def normalize_size(raw: object) -> Optional[str]:
    s = normalize(raw).replace(" ", "")
    if s == "XXL":
        s = "2XL"
    elif s == "XXXL":
        s = "3XL"
    return s if s in VALID_SIZES else None


def parse_name(name: object) -> Optional[dict]:
    """
    Parsea un nombre completo de Quenti.
    Devuelve dict con tipo/base/calidad/acabado/talla, o None si no es válido.
    """
    n = normalize(name)
    if not n:
        return None

    mh = _PREFIX_RE.match(n)
    if not mh:
        return None
    tipo = TIPO_BUSO if mh.group("head") == "BUSO" else TIPO_CAMISETA
    rest = mh.group("rest")

    msz = _SIZE_TAIL_RE.match(rest)
    if not msz:
        return None
    body = msz.group("body").strip()
    size = normalize_size(msz.group("size"))
    if size is None:
        return None

    acabado = ACABADO_SIN
    if _PARCHES_CON_RE.search(body):
        acabado = ACABADO_CON
        body = re.sub(r"\s+", " ", _PARCHES_CON_RE.sub("", body)).strip()
    if _PARCHES_SIN_RE.search(body):
        acabado = ACABADO_SIN
        body = re.sub(r"\s+", " ", _PARCHES_SIN_RE.sub("", body)).strip()

    calidad = None
    for code, rx in _CALIDAD_PATTERNS:
        if rx.search(body):
            calidad = code
            body = re.sub(r"\s+", " ", rx.sub("", body)).strip()
            break

    base = body.strip()
    if not base:
        return None

    return {
        "tipo":    tipo,
        "base":    base,
        "calidad": calidad,
        "acabado": acabado,
        "talla":   size,
    }


def parse_row(row: dict) -> Optional[dict]:
    """
    Parsea una fila del Excel (dict con encabezados de Quenti).
    Devuelve dict completo con sku/stock/precio/raw_name o None si no aplica.
    """
    name = row.get("nombre")
    parsed = parse_name(name)
    if parsed is None:
        return None
    return {
        **parsed,
        "sku":      str(row.get("codigo_barras") or "").strip(),
        "stock":    int(row.get("existencias") or 0),
        "precio":   int(row.get("precioVenta_con_impuesto") or 0),
        "raw_name": str(name or ""),
    }


def read_xlsx(path: str) -> Iterator[dict]:
    """
    Generador que entrega filas como dicts asociativos por encabezado de columna.
    Usa openpyxl en modo read_only para soportar archivos grandes.
    """
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    headers = None
    for raw in ws.iter_rows(values_only=True):
        if headers is None:
            headers = [str(h or "").strip() for h in raw]
            continue
        yield {h: v for h, v in zip(headers, raw)}


def parse_xlsx(path: str) -> list[dict]:
    """Lee + parsea el .xlsx entero. Devuelve solo filas válidas."""
    return [r for r in (parse_row(row) for row in read_xlsx(path)) if r]


def group_by_family(rows: Iterable[dict]) -> dict:
    """Agrupa filas parseadas por (tipo, base) → lista de variaciones."""
    out: dict[tuple[str, str], dict] = {}
    for r in rows:
        key = (r["tipo"], r["base"])
        if key not in out:
            out[key] = {"tipo": r["tipo"], "base": r["base"], "variaciones": []}
        out[key]["variaciones"].append({
            "calidad": r["calidad"],
            "acabado": r["acabado"],
            "talla":   r["talla"],
            "sku":     r["sku"],
            "stock":   r["stock"],
            "precio":  r["precio"],
        })
    return out


def detect_new_sizes(rows: Iterable[dict]) -> set[str]:
    """Devuelve las tallas detectadas que aún no existen en WooCommerce."""
    seen = {r["talla"] for r in rows}
    return seen & set(SIZES_NOT_YET_IN_WC)
