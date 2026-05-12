import openpyxl
import sys
import os

os.chdir(r'C:\Proyectos\b370\PROTOCOLO 1.1 B370\B370-MANAGER')
wb = openpyxl.load_workbook('data/CUENTI INVENTARIO 6 ABRIL.xlsx')
ws = wb.active
headers = [cell.value for cell in ws[1]]

print('ARSENAL RETRO 1,1 - todas las tallas:')
for row in ws.iter_rows(min_row=2, values_only=True):
    nombre = str(row[1]).upper() if row[1] else ''
    if 'ARSENAL RETRO 1' in nombre and 'HOMBRE' in nombre:
        d = dict(zip(headers, row))
        print(f"  nombre: {d['nombre']}")
        print(f"  codigo_barras: {d['codigo_barras']}")
        print(f"  referencia: {d['referencia']}")
        print()
