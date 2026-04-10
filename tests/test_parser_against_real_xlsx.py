"""
Valida el parser de class-quenti.php replicando su lógica exacta en Python
y corriéndolo contra el archivo real CUENTI INVENTARIO 6 ABRIL.xlsx.

No reemplaza a los tests PHP — es una sanity check ejecutable sin servidor.
La lógica debe mantenerse 1:1 con B370_Manager_Quenti::parse_name().
"""
import re
import sys
import io
import os
from pathlib import Path
from collections import Counter, defaultdict
from openpyxl import load_workbook
from dotenv import load_dotenv

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

# Carga variables desde .env (busca en la raíz del proyecto)
load_dotenv(Path(__file__).parent.parent / '.env')

XLSX = os.environ.get('XLSX_PATH', r'C:\Users\USUARIO\Downloads\CUENTI INVENTARIO 6 ABRIL.xlsx')
WC_URL = os.environ.get('WC_URL', '')
WC_CK  = os.environ.get('WC_CK', '')
WC_CS  = os.environ.get('WC_CS', '')

VALID_SIZES = {'XS', 'S', 'M', 'L', 'XL', '2XL', '3XL', '4XL', '5XL', '6XL'}

def normalize(s):
    if s is None:
        return ''
    s = str(s)
    s = s.lstrip('\ufeff')
    s = s.upper()
    trans = str.maketrans('ÁÉÍÓÚÜÑ', 'AEIOUUN')
    s = s.translate(trans)
    s = re.sub(r'\s+', ' ', s).strip()
    return s

def normalize_size(raw):
    s = normalize(raw).replace(' ', '')
    if s == 'XXL':
        s = '2XL'
    if s == 'XXXL':
        s = '3XL'
    return s if s in VALID_SIZES else None

def parse_name(name):
    n = normalize(name)
    if not n:
        return None
    mh = re.match(r'^(?P<head>CAMISETA|BUSO)(?:\s+DEL?)?\s+(?P<rest>.+)$', n)
    if not mh:
        return None
    tipo = 'buso' if mh.group('head') == 'BUSO' else 'camiseta'
    rest = mh.group('rest')

    m = re.match(r'^(?P<body>.+?)\s*/\s*(?P<size>[A-Z0-9]+)\s*$', rest)
    if not m:
        return None
    body = m.group('body').strip()
    size = normalize_size(m.group('size'))
    if size is None:
        return None

    acabado = 'sin_parches'
    if re.search(r'\bCON\s+PARCHES\b', body):
        acabado = 'con_parches'
        body = re.sub(r'\bCON\s+PARCHES\b', '', body)
        body = re.sub(r'\s+', ' ', body).strip()
    if re.search(r'\bSIN\s+PARCHES\b', body):
        acabado = 'sin_parches'
        body = re.sub(r'\bSIN\s+PARCHES\b', '', body)
        body = re.sub(r'\s+', ' ', body).strip()

    calidad = None
    patterns = [
        ('tipo_original', r'\bTIPO\s+ORIGINAL\b'),
        ('1.1',           r'(?<![\d\.,])1[.,]1(?![\d])'),
        ('version_fan',   r'\bFAN\b'),
    ]
    for code, rx in patterns:
        if re.search(rx, body):
            calidad = code
            body = re.sub(rx, '', body)
            body = re.sub(r'\s+', ' ', body).strip()
            break

    base = body.strip()
    if not base:
        return None

    return {'tipo': tipo, 'base': base, 'calidad': calidad, 'acabado': acabado, 'talla': size}


def main():
    wb = load_workbook(XLSX, data_only=True, read_only=True)
    ws = wb['Hoja1']
    total = 0
    parsed_ok = 0
    not_matched_camiseta_buso = 0
    rejected_size = 0
    rejected_no_slash = 0
    by_tipo = Counter()
    by_calidad = Counter()
    by_acabado = Counter()
    by_size = Counter()
    families = defaultdict(list)
    anomalies_con_parches = []
    sample_fail_camiseta = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        name = row[1]
        if not isinstance(name, str):
            continue
        total += 1
        u = name.upper().strip()
        is_target = u.startswith('CAMISETA') or u.startswith('BUSO')
        if not is_target:
            not_matched_camiseta_buso += 1
            continue

        p = parse_name(name)
        if p is None:
            # Investigar por qué falló
            if '/' not in u:
                rejected_no_slash += 1
            else:
                size_raw = u.rsplit('/', 1)[1].strip()
                if normalize_size(size_raw) is None:
                    rejected_size += 1
                else:
                    if len(sample_fail_camiseta) < 10:
                        sample_fail_camiseta.append(name)
            continue

        parsed_ok += 1
        by_tipo[p['tipo']] += 1
        by_calidad[p['calidad']] += 1
        by_acabado[p['acabado']] += 1
        by_size[p['talla']] += 1
        families[(p['tipo'], p['base'])].append(p)
        if p['acabado'] == 'con_parches' and p['calidad'] != 'tipo_original':
            anomalies_con_parches.append(name)

    print(f'Filas totales con nombre:          {total}')
    print(f'  No son CAMISETA/BUSO (ignoradas): {not_matched_camiseta_buso}')
    print(f'  CAMISETA/BUSO rechazadas:')
    print(f'    - sin "/" final (sin talla):    {rejected_no_slash}')
    print(f'    - talla no soportada:           {rejected_size}')
    print(f'  CAMISETA/BUSO parseadas OK:       {parsed_ok}')
    print()
    print('Por tipo:      ', dict(by_tipo))
    print('Por calidad:   ', dict(by_calidad))
    print('Por acabado:   ', dict(by_acabado))
    print('Por talla:     ', dict(by_size))
    print(f'Familias únicas (tipo+base): {len(families)}')
    print()
    print('Top 10 familias por nº de variaciones:')
    top = sorted(families.items(), key=lambda x: -len(x[1]))[:10]
    for (tipo, base), vs in top:
        print(f'  [{tipo:8}] {base}  → {len(vs)} variaciones')
    print()
    print(f'Anomalías "con parches" sin "tipo original": {len(anomalies_con_parches)}')
    for n in anomalies_con_parches[:5]:
        print(f'  ! {n}')
    print()
    if sample_fail_camiseta:
        print('Muestras de CAMISETA/BUSO que NO parsearon pese a tener talla válida:')
        for n in sample_fail_camiseta:
            print(f'  ? {n}')

    # Sanity check: el parser debe cubrir al menos las ~1090 CAMISETA/BUSO
    # detectadas en la exploración inicial.
    assert parsed_ok >= 1000, f'Parser coverage too low: {parsed_ok} < 1000'
    print('\nOK — cobertura del parser aceptable.')

if __name__ == '__main__':
    main()
